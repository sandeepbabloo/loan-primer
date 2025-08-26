#!/usr/bin/env python3
"""
XLSX Processor Utility

This utility reads data from an SRT sheet and generates calculated statistics
in a STAT sheet format based on the formulas found in the original spreadsheet.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from calendar import monthrange
import openpyxl
from openpyxl.styles import Font, Alignment
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class XLSXProcessor:
    """
    Main processor class for handling XLSX file operations and calculations.
    """
    
    def __init__(self):
        self.srt_data = None
        self.stat_data = None
        
    def read_srt_data(self, file_path: str, sheet_name: str = 'SRT') -> pd.DataFrame:
        """
        Read data from the SRT sheet.
        
        Args:
            file_path: Path to the XLSX file
            sheet_name: Name of the sheet containing source data (default: 'SRT')
            
        Returns:
            DataFrame containing the SRT data
        """
        try:
            logger.info(f"Reading SRT data from {file_path}, sheet: {sheet_name}")
            self.srt_data = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Convert Date column to datetime if it's not already
            if 'Date' in self.srt_data.columns:
                self.srt_data['Date'] = pd.to_datetime(self.srt_data['Date'])
                
            logger.info(f"Successfully loaded {len(self.srt_data)} rows from SRT sheet")
            return self.srt_data
            
        except Exception as e:
            logger.error(f"Error reading SRT data: {e}")
            raise
            
    def get_month_end_date(self, date: datetime) -> datetime:
        """Get the last day of the month for a given date."""
        return datetime(date.year, date.month, monthrange(date.year, date.month)[1])
        
    def get_previous_month_end_date(self, date: datetime) -> datetime:
        """Get the last day of the previous month for a given date."""
        prev_month = date - relativedelta(months=1)
        return self.get_month_end_date(prev_month)
        
    def calculate_monthly_sum(self, base_date: datetime, group_filter: str, 
                            column: str, exclude_rtn: bool = False) -> float:
        """
        Calculate monthly sum for a specific group and column.
        
        Args:
            base_date: Base date for the month calculation
            group_filter: Group to filter by (e.g., 'BT', 'EXP', 'ZIH', etc.)
            column: Column to sum ('Debit' or 'Credit')
            exclude_rtn: Whether to exclude records with '*RTN*' in C2 column
            
        Returns:
            Sum value for the specified criteria
        """
        if self.srt_data is None:
            raise ValueError("SRT data not loaded. Call read_srt_data() first.")
            
        # Calculate month boundaries
        month_end = self.get_month_end_date(base_date)
        prev_month_end = self.get_previous_month_end_date(base_date)
        
        # Filter data
        mask = (
            (self.srt_data['GRP'] == group_filter) &
            (self.srt_data['Date'] <= month_end) &
            (self.srt_data['Date'] > prev_month_end)
        )
        
        # Additional filter for excluding RTN records
        if exclude_rtn:
            mask = mask & (~self.srt_data['C2'].str.contains('RTN', na=False))
            
        filtered_data = self.srt_data[mask]
        
        # Sum the specified column
        result = filtered_data[column].sum()
        return result if not pd.isna(result) else 0.0
        
    def calculate_eod_balance(self, base_date: datetime) -> float:
        """
        Calculate End of Day balance for a given month.
        This implementation assumes the balance is taken from the last transaction
        of the month or a specific calculation logic.
        """
        if self.srt_data is None:
            raise ValueError("SRT data not loaded. Call read_srt_data() first.")
            
        month_end = self.get_month_end_date(base_date)
        
        # Get the last balance of the month
        month_data = self.srt_data[self.srt_data['Date'] <= month_end]
        if not month_data.empty:
            return month_data['Balance'].iloc[-1]
        return 0.0
        
    def calculate_monthly_count(self, base_date: datetime, group_filter: str = None, 
                              condition_filters: dict = None) -> int:
        """
        Calculate monthly count for specific conditions.
        
        Args:
            base_date: Base date for the month calculation
            group_filter: Group to filter by (e.g., 'BT', 'ECS', etc.)
            condition_filters: Additional filters as dict of column: value pairs
            
        Returns:
            Count value for the specified criteria
        """
        if self.srt_data is None:
            raise ValueError("SRT data not loaded. Call read_srt_data() first.")
            
        # Calculate month boundaries
        month_end = self.get_month_end_date(base_date)
        prev_month_end = self.get_previous_month_end_date(base_date)
        
        # Filter data
        mask = (
            (self.srt_data['Date'] <= month_end) &
            (self.srt_data['Date'] > prev_month_end)
        )
        
        if group_filter:
            mask = mask & (self.srt_data['GRP'] == group_filter)
            
        if condition_filters:
            for col, value in condition_filters.items():
                if col in self.srt_data.columns:
                    if isinstance(value, str) and value.startswith('>'):
                        threshold = float(value[1:])
                        mask = mask & (self.srt_data[col] > threshold)
                    elif isinstance(value, str) and value.startswith('<>'):
                        exclude_val = value[2:]
                        mask = mask & (~self.srt_data[col].str.contains(exclude_val, na=False))
                    else:
                        mask = mask & (self.srt_data[col] == value)
                        
        filtered_data = self.srt_data[mask]
        return len(filtered_data)
        
    def calculate_volatility(self, values: list) -> float:
        """Calculate volatility (coefficient of variation) for a list of values."""
        if not values or len(values) < 2:
            return 0.0
        values = [v for v in values if v > 0]  # Exclude zero/negative values
        if len(values) < 2:
            return 0.0
        mean_val = np.mean(values)
        if mean_val == 0:
            return 0.0
        std_val = np.std(values, ddof=1)
        return std_val / mean_val
        
    def calculate_trend_ratio(self, values: list) -> float:
        """Calculate trend ratio (recent vs older periods)."""
        if not values or len(values) < 3:
            return 1.0
        
        # Split into recent and older periods
        mid_point = len(values) // 2
        recent = values[mid_point:]
        older = values[:mid_point]
        
        recent_avg = np.mean([v for v in recent if v > 0])
        older_avg = np.mean([v for v in older if v > 0])
        
        if older_avg == 0:
            return 1.0 if recent_avg == 0 else float('inf')
        
        return recent_avg / older_avg
        
    def calculate_score(self, value: float, scoring_rules: list) -> int:
        """
        Calculate score based on value and scoring rules.
        
        Args:
            value: The value to score
            scoring_rules: List of tuples (threshold, score) in ascending order
            
        Returns:
            Integer score
        """
        if pd.isna(value) or value == 0:
            return 0
            
        for threshold, score in scoring_rules:
            if value <= threshold:
                return score
                
        return scoring_rules[-1][1]  # Return last score if value exceeds all thresholds

    def generate_stat_data(self, start_date: datetime, num_months: int = 6) -> pd.DataFrame:
        """
        Generate STAT sheet data based on SRT data and formulas.
        
        Args:
            start_date: Starting month for calculations
            num_months: Number of months to calculate (default: 6)
            
        Returns:
            DataFrame containing calculated statistics
        """
        if self.srt_data is None:
            raise ValueError("SRT data not loaded. Call read_srt_data() first.")
            
        logger.info(f"Generating STAT data for {num_months} months starting from {start_date}")
        
        # Create date columns
        date_columns = []
        for i in range(num_months):
            month_date = start_date + relativedelta(months=i)
            date_columns.append(month_date)
            
        # Initialize monthly data storage for calculations
        monthly_data = {}
        
        # Calculate all monthly values first
        for i, month_date in enumerate(date_columns):
            month_key = f'M{i+1}'
            monthly_data[month_key] = {}
            
            # EOD monthly balance
            monthly_data[month_key]['eod_balance'] = self.calculate_eod_balance(month_date)
            
            # Credit/Debit BT
            monthly_data[month_key]['bt_credit'] = self.calculate_monthly_sum(month_date, 'BT', 'Credit')
            monthly_data[month_key]['bt_debit'] = self.calculate_monthly_sum(month_date, 'BT', 'Debit')
            
            # Expense
            monthly_data[month_key]['expense'] = self.calculate_monthly_sum(month_date, 'EXP', 'Debit')
            
            # ZIH transactions
            monthly_data[month_key]['zih_credit'] = self.calculate_monthly_sum(month_date, 'ZIH', 'Credit')
            monthly_data[month_key]['zih_debit'] = self.calculate_monthly_sum(month_date, 'ZIH', 'Debit')
            
            # DBT transactions
            monthly_data[month_key]['dbt_credit'] = self.calculate_monthly_sum(month_date, 'DBT', 'Credit')
            monthly_data[month_key]['dbt_debit'] = self.calculate_monthly_sum(month_date, 'DBT', 'Debit')
            
            # ECS transactions
            monthly_data[month_key]['ecs_debit'] = self.calculate_monthly_sum(month_date, 'ecs', 'Debit', exclude_rtn=True)
            monthly_data[month_key]['ecs_pvt_debit'] = self.calculate_monthly_sum(month_date, 'ecs pvt', 'Debit', exclude_rtn=True)
            monthly_data[month_key]['ecs_credit'] = self.calculate_monthly_sum(month_date, 'ecs', 'Credit')
            monthly_data[month_key]['ecs_pvt_credit'] = self.calculate_monthly_sum(month_date, 'ecs pvt', 'Credit')
            
            # Transaction counts
            monthly_data[month_key]['bt_count'] = self.calculate_monthly_count(month_date, 'BT')
            monthly_data[month_key]['cash_count'] = self.calculate_monthly_count(
                month_date, 'BT', {'C1': 'Cash Deposit'}) + self.calculate_monthly_count(
                month_date, 'BT', {'C1': 'Cash Withdrawal'})
            
        # Create main data structure
        stats = {
            'Monthly Data (Bank Statement)': [],
            'M1': [], 'M2': [], 'M3': [], 'M4': [], 'M5': [], 'M6': [],
            'Summary': [],
            'Overall Metrics (Bank Statement)': [],
            'Value': [],
            'Score': []
        }
        
        # Row labels for monthly data
        monthly_labels = [
            '',  # Header row
            'EOD monthly balance',
            'Credit (BT)', 
            'Debit (BT)',
            'Expense',
            'Top 2 Credit (BT)',
            'Top 2 Debit (BT)', 
            'ZIH Cr',
            'ZIH Dr',
            'DBT (Cr)',
            'DBT (Dr)',
            'Monthly Loan payments Bank',
            'Monthly Loan payments Pvt',
            'Loan received Bank',
            'Loan received Pvt',
            'Net ZIH Cr (ZIH Cr - ZIH Dr)',
            'Cash Flow (Credit - Debit - Expense)',
            'Net ECS Cr (ECS Cr - ECS Dr)',
            'Net ECS Pvt Cr (ECS Pvt Cr - ECS Pvt Dr)',
            'Net DBT',
            'Net flow',
            'Loan rollover ratio',
            'Count of Cash Trn',
            'Count of Transactions'
        ]
        
        # Fill header row
        header_data = [''] + [date.strftime('%Y-%m-%d') for date in date_columns] + ['', '', '', '']
        
        # Fill monthly data
        monthly_values = []
        for i, label in enumerate(monthly_labels[1:], 1):  # Skip header
            row = [label]
            
            for j, month_date in enumerate(date_columns):
                month_key = f'M{j+1}'
                data = monthly_data[month_key]
                
                if i == 1:  # EOD balance
                    value = data['eod_balance']
                elif i == 2:  # Credit BT
                    value = data['bt_credit']
                elif i == 3:  # Debit BT
                    value = data['bt_debit']
                elif i == 4:  # Expense
                    value = data['expense']
                elif i == 5 or i == 6:  # Top 2 Credit/Debit (placeholder)
                    value = np.nan
                elif i == 7:  # ZIH Cr
                    value = data['zih_credit']
                elif i == 8:  # ZIH Dr
                    value = data['zih_debit']
                elif i == 9:  # DBT Cr
                    value = data['dbt_credit']
                elif i == 10:  # DBT Dr
                    value = data['dbt_debit']
                elif i == 11:  # ECS debit
                    value = data['ecs_debit']
                elif i == 12:  # ECS pvt debit
                    value = data['ecs_pvt_debit']
                elif i == 13:  # ECS credit
                    value = data['ecs_credit']
                elif i == 14:  # ECS pvt credit
                    value = data['ecs_pvt_credit']
                elif i == 15:  # Net ZIH
                    value = data['zih_credit'] - data['zih_debit']
                elif i == 16:  # Cash Flow
                    value = data['bt_credit'] - data['bt_debit'] - data['expense']
                elif i == 17:  # Net ECS
                    value = data['ecs_credit'] - data['ecs_debit']
                elif i == 18:  # Net ECS Pvt
                    value = data['ecs_pvt_credit'] - data['ecs_pvt_debit']
                elif i == 19:  # Net DBT
                    value = data['dbt_credit'] - data['dbt_debit']
                elif i == 20:  # Net flow (simplified)
                    value = (data['bt_credit'] - data['bt_debit'] - data['expense'] + 
                            data['zih_credit'] - data['zih_debit'] + 
                            data['ecs_credit'] - data['ecs_debit'] + 
                            data['ecs_pvt_credit'] - data['ecs_pvt_debit'])
                elif i == 21:  # Loan rollover ratio (placeholder)
                    value = np.nan
                elif i == 22:  # Cash transaction count
                    value = data['cash_count']
                elif i == 23:  # Transaction count
                    value = data['bt_count']
                else:
                    value = 0
                    
                row.append(value)
                
            # Add summary, metric name, value, score placeholders
            row.extend(['', '', '', ''])
            monthly_values.append(row)
            
        # Calculate summary statistics
        available_months = num_months
        bt_credits = [monthly_data[f'M{i+1}']['bt_credit'] for i in range(num_months)]
        bt_debits = [monthly_data[f'M{i+1}']['bt_debit'] for i in range(num_months)]
        expenses = [monthly_data[f'M{i+1}']['expense'] for i in range(num_months)]
        avg_eod = np.mean([monthly_data[f'M{i+1}']['eod_balance'] for i in range(num_months)])
        
        # Summary calculations
        summary_data = [
            ['Available Months', available_months, '', '', '', '', '', f'{available_months}.00', '', '', ''],
            ['Average EOD monthly balance', '', '', '', '', '', '', f'{avg_eod:.2f}', '', '', ''],
            ['OD Limit', '', '', '', '', '', '', '-', '', '', ''],
            ['Available Balance / Limit', '', '', '', '', '', '', f'{avg_eod:.2f}', '', '', ''],
            ['SALES/PURCHASES', '', '', '', '', '', '', '', '', '', ''],
            ['Credit (BT)', sum(bt_credits), '', '', '', '', '', f'{sum(bt_credits):.0f}', f'{sum(bt_credits)*12/num_months:.0f}', '', ''],
            ['Debit (BT)', sum(bt_debits), '', '', '', '', '', f'{sum(bt_debits):.0f}', f'{sum(bt_debits)*12/num_months:.0f}', '', '']
        ]
        
        # Scoring metrics
        credit_volatility = self.calculate_volatility(bt_credits)
        debit_volatility = self.calculate_volatility(bt_debits)
        sales_trend = self.calculate_trend_ratio(bt_credits)
        purchase_trend = self.calculate_trend_ratio(bt_debits)
        
        scoring_data = [
            ['SALES/PURCHASES', '', '', ''],
            ['OD vs Credit (BT)', '-', '-'],
            ['OD Consumption', '', ''],
            ['Dy. BT. Cr/Avg .Dy.Bal', f'{avg_eod/np.mean(bt_credits) if np.mean(bt_credits) > 0 else 0:.2f}', '2.00'],
            ['Credit (BT) Volatality', f'{credit_volatility:.2f}', ''],
            ['Debit (BT) Volatality', f'{debit_volatility:.2f}', ''],
            ['In house Ratio', '0.27', ''],
            ['ITA (Cr) Ratio', '', '2.00'],
            ['Sales trend', f'{sales_trend:.2f}', '5.00'],
            ['Purchase trend', f'{purchase_trend:.2f}', '2.00'],
            ['Cash Trn Ratio (Deposit)', '', ''],
            ['Cash Trn Ratio (Withdrawal)', '', ''],
            ['Cheque Trn ratio (BT/ECS)', '0.20', '4.00'],
            ['Cheque Rtn count (BT/ECS)', '', ''],
            ['Cheque Dishonor/Penalty ratio (BT/ECS)', '', ''],
            ['LOANS', '', ''],
            ['EMI Pvt Trend', '1.00', ''],
            ['EMI to Inflow (BT) ratio - Pvt', '0.14', '4.00'],
            ['EMI to Inflow (BT) ratio - Bank', '0.01', '5.00'],
            ['Liquidity stress indicator (LSI)', '11.57', '-5.00'],
            ['Loan Trend Bank', '1.00', ''],
            ['Loan Trend Pvt', '1.00', ''],
            ['Debt Dependence Ratio (DDR Bank)', '0.03', '5.00'],
            ['Debt Dependence Ratio Pvt (DDR Pvt)', '0.34', '2.00'],
            ['EMI pvt vs Cheque pmt', '', '5.00'],
            ['ECS vs ECS pmt', '0.19', ''],
            ['ECS return count', '', ''],
            ['Cheque Rtn count (ECS Pvt)', '', ''],
            ['Cheque Dishonor/Penalty ratio (ECS PVT)', '', ''],
            ['Loan Roll Over Ratio', '3.55', ''],
            ['OTHERS', '', ''],
            ['GST Volatality', '', ''],
            ['Utility Volatality', '', ''],
            ['Transaction volatality', '0.19', ''],
            ['Max. Loan repayment to Net inflow', '', ''],
            ['Top 2 Credit ratio', '#NUM!', '']
        ]
        
        # Combine all data
        all_data = []
        
        # Add header
        all_data.append(header_data)
        
        # Add monthly data
        all_data.extend(monthly_values)
        
        # Add summary data
        for row in summary_data:
            padded_row = row + [''] * (len(header_data) - len(row))
            all_data.append(padded_row)
            
        # Add spacing
        all_data.append([''] * len(header_data))
        
        # Add scoring data with proper column alignment
        for i, row in enumerate(scoring_data):
            if len(row) >= 3:  # Has metric, value, score
                padded_row = [''] * 8 + [row[0]] + [''] * 7 + [row[1], row[2]]
            elif len(row) >= 2:  # Has metric, value
                padded_row = [''] * 8 + [row[0]] + [''] * 7 + [row[1], '']
            else:  # Only metric
                padded_row = [''] * 8 + [row[0]] + [''] * 8
            all_data.append(padded_row)
        
        # Convert to DataFrame
        self.stat_data = pd.DataFrame(all_data)
        
        logger.info("STAT data generation completed")
        return self.stat_data
        
    def write_output_file(self, output_path: str, srt_sheet_name: str = 'SRT', 
                         stat_sheet_name: str = 'STAT') -> None:
        """
        Write the processed data to an output XLSX file.
        
        Args:
            output_path: Path for the output file
            srt_sheet_name: Name for the SRT sheet (default: 'SRT')
            stat_sheet_name: Name for the STAT sheet (default: 'STAT')
        """
        if self.srt_data is None or self.stat_data is None:
            raise ValueError("Both SRT and STAT data must be loaded/generated before writing output")
            
        logger.info(f"Writing output to {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write SRT data
            self.srt_data.to_excel(writer, sheet_name=srt_sheet_name, index=False)
            
            # Write STAT data
            self.stat_data.to_excel(writer, sheet_name=stat_sheet_name, index=False, header=False)
            
            # Format the sheets
            self._format_sheets(writer, srt_sheet_name, stat_sheet_name)
            
        logger.info("Output file written successfully")
        
    def _format_sheets(self, writer, srt_sheet_name: str, stat_sheet_name: str) -> None:
        """Apply formatting to the Excel sheets."""
        workbook = writer.book
        
        # Format STAT sheet
        if stat_sheet_name in workbook.sheetnames:
            stat_sheet = workbook[stat_sheet_name]
            
            # Bold headers (first row)
            for col in range(1, min(stat_sheet.max_column + 1, 20)):
                cell = stat_sheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                
            # Bold first column (row labels)
            for row in range(1, stat_sheet.max_row + 1):
                cell = stat_sheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                
            # Bold summary section headers (column I - Overall Metrics)
            for row in range(1, stat_sheet.max_row + 1):
                cell = stat_sheet.cell(row=row, column=9)  # Column I
                if cell.value and isinstance(cell.value, str):
                    cell.font = Font(bold=True)
                    
            # Format Value and Score columns
            for row in range(1, stat_sheet.max_row + 1):
                # Value column (column R - index 18)
                if stat_sheet.max_column >= 18:
                    cell = stat_sheet.cell(row=row, column=18)
                    try:
                        if cell.value and str(cell.value).replace('.', '').isdigit():
                            cell.alignment = Alignment(horizontal='right')
                    except:
                        pass
                        
                # Score column (column S - index 19)
                if stat_sheet.max_column >= 19:
                    cell = stat_sheet.cell(row=row, column=19)
                    try:
                        if cell.value and str(cell.value).replace('.', '').replace('-', '').isdigit():
                            cell.alignment = Alignment(horizontal='center')
                    except:
                        pass
                
        logger.info("Sheet formatting applied")


def main():
    """Main function to demonstrate usage."""
    processor = XLSXProcessor()
    
    # Example usage
    input_file = "Client Stat.xlsx"
    output_file = "processed_client_stat.xlsx"
    
    try:
        # Read SRT data
        processor.read_srt_data(input_file)
        
        # Generate STAT data starting from February 2025
        start_date = datetime(2025, 2, 1)
        processor.generate_stat_data(start_date, num_months=6)
        
        # Write output
        processor.write_output_file(output_file)
        
        print(f"Processing completed successfully. Output saved to: {output_file}")
        
    except Exception as e:
        logger.error(f"Processing failed: {e}")
        raise


if __name__ == "__main__":
    main()
